import pandas as pd
from typing import Dict, List, Any, Optional
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np

class TableProcessor:

    def __init__(self):
        self.table_analysis_config = {'min_rows_for_table': 2, 'min_cols_for_table': 2, 'max_empty_cells_percentage': 0.8, 'header_detection_threshold': 0.7}

    def process_xlsx_tables(self, file_path: str) -> Dict[str, Any]:
        try:
            workbook = load_workbook(file_path, data_only=True)
            processed_tables = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_tables = self._extract_tables_from_sheet(sheet, sheet_name)
                if sheet_tables:
                    processed_tables[sheet_name] = sheet_tables
            return {'file_path': file_path, 'sheets': processed_tables, 'total_sheets': len(workbook.sheetnames), 'sheets_with_tables': len(processed_tables)}
        except Exception as e:
            return {'error': f'Error processing XLSX file: {str(e)}', 'file_path': file_path}

    def _extract_tables_from_sheet(self, sheet, sheet_name: str) -> List[Dict]:
        tables = []
        try:
            max_row = sheet.max_row
            max_col = sheet.max_column
            if max_row < 2 or max_col < 2:
                return tables
            data = []
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    value = cell.value
                    if value is None:
                        value = ''
                    elif isinstance(value, (int, float)):
                        value = str(value)
                    elif isinstance(value, str):
                        value = value.strip()
                    row_data.append(value)
                data.append(row_data)
            df = pd.DataFrame(data)
            table_regions = self._find_table_regions(df)
            for i, region in enumerate(table_regions):
                table_data = self._extract_table_from_region(df, region, sheet_name, i)
                if table_data:
                    tables.append(table_data)
        except Exception as e:
            print(f'Error extracting tables from sheet {sheet_name}: {e}')
        return tables

    def _find_table_regions(self, df: pd.DataFrame) -> List[Dict]:
        regions = []
        rows, cols = df.shape
        non_empty_rows = []
        for i in range(rows):
            if not df.iloc[i].isna().all() and (not (df.iloc[i] == '').all()):
                non_empty_rows.append(i)
        if len(non_empty_rows) < 2:
            return regions
        current_start = non_empty_rows[0]
        current_end = non_empty_rows[0]
        for i in range(1, len(non_empty_rows)):
            if non_empty_rows[i] == current_end + 1:
                current_end = non_empty_rows[i]
            else:
                if current_end - current_start >= 1:
                    regions.append({'start_row': current_start, 'end_row': current_end, 'start_col': 0, 'end_col': cols - 1})
                current_start = non_empty_rows[i]
                current_end = non_empty_rows[i]
        if current_end - current_start >= 1:
            regions.append({'start_row': current_start, 'end_row': current_end, 'start_col': 0, 'end_col': cols - 1})
        return regions

    def _extract_table_from_region(self, df: pd.DataFrame, region: Dict, sheet_name: str, table_index: int) -> Optional[Dict]:
        try:
            start_row, end_row = (region['start_row'], region['end_row'])
            start_col, end_col = (region['start_col'], region['end_col'])
            table_df = df.iloc[start_row:end_row + 1, start_col:end_col + 1].copy()
            table_df = self._clean_table_data(table_df)
            if table_df.empty or table_df.shape[0] < 2 or table_df.shape[1] < 2:
                return None
            has_headers = self._detect_headers(table_df)
            table_data = {'table_index': table_index, 'sheet_name': sheet_name, 'region': region, 'dimensions': {'rows': table_df.shape[0], 'columns': table_df.shape[1]}, 'has_headers': has_headers, 'data': table_df.values.tolist(), 'text': self._table_to_text(table_df, has_headers), 'analysis': self._analyze_table_structure(table_df)}
            return table_data
        except Exception as e:
            print(f'Error extracting table from region: {e}')
            return None

    def _clean_table_data(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        df = df.fillna('')
        row_empty_ratio = (df == '').sum(axis=1) / df.shape[1]
        col_empty_ratio = (df == '').sum(axis=0) / df.shape[0]
        df = df[row_empty_ratio < self.table_analysis_config['max_empty_cells_percentage']]
        df = df.loc[:, col_empty_ratio < self.table_analysis_config['max_empty_cells_percentage']]
        return df

    def _detect_headers(self, df: pd.DataFrame) -> bool:
        if df.shape[0] < 2:
            return False
        first_row = df.iloc[0]
        other_rows = df.iloc[1:]
        first_row_non_empty = (first_row != '').sum()
        other_rows_avg_non_empty = (other_rows != '').sum(axis=1).mean()
        return first_row_non_empty > other_rows_avg_non_empty * self.table_analysis_config['header_detection_threshold']

    def _table_to_text(self, df: pd.DataFrame, has_headers: bool) -> str:
        text_lines = []
        if has_headers and df.shape[0] > 0:
            headers = df.iloc[0]
            text_lines.append('Headers: ' + '\t'.join((str(h) for h in headers)))
            text_lines.append('-' * 50)
            for i in range(1, df.shape[0]):
                row = df.iloc[i]
                text_lines.append('\t'.join((str(cell) for cell in row)))
        else:
            for i in range(df.shape[0]):
                row = df.iloc[i]
                text_lines.append('\t'.join((str(cell) for cell in row)))
        return '\n'.join(text_lines)

    def _analyze_table_structure(self, df: pd.DataFrame) -> Dict:
        analysis = {'total_cells': df.shape[0] * df.shape[1], 'non_empty_cells': (df != '').sum().sum(), 'empty_cells': (df == '').sum().sum(), 'data_types': {}, 'numeric_columns': [], 'text_columns': [], 'date_columns': []}
        for col_idx in range(df.shape[1]):
            col_data = df.iloc[:, col_idx]
            non_empty_data = col_data[col_data != '']
            if len(non_empty_data) == 0:
                continue
            numeric_count = 0
            for value in non_empty_data:
                try:
                    float(str(value))
                    numeric_count += 1
                except (ValueError, TypeError):
                    pass
            numeric_ratio = numeric_count / len(non_empty_data)
            if numeric_ratio > 0.5:
                analysis['numeric_columns'].append(col_idx)
            else:
                analysis['text_columns'].append(col_idx)
        analysis['data_types']['numeric_columns'] = len(analysis['numeric_columns'])
        analysis['data_types']['text_columns'] = len(analysis['text_columns'])
        return analysis

    def get_table_summary(self, processed_tables: Dict) -> str:
        summary_lines = []
        for sheet_name, tables in processed_tables.items():
            summary_lines.append(f'Sheet: {sheet_name}')
            summary_lines.append(f'  Tables found: {len(tables)}')
            for table in tables:
                dims = table['dimensions']
                analysis = table['analysis']
                summary_lines.append(f"    Table {table['table_index']}: {dims['rows']}x{dims['columns']} ({analysis['non_empty_cells']}/{analysis['total_cells']} cells filled)")
            summary_lines.append('')
        return '\n'.join(summary_lines)

    def export_tables_to_json(self, processed_tables: Dict, output_path: str) -> bool:
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(processed_tables, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            print(f'Error exporting tables to JSON: {e}')
            return False

def process_xlsx_file(file_path: str) -> Dict[str, Any]:
    processor = TableProcessor()
    return processor.process_xlsx_tables(file_path)

def get_table_summary(processed_tables: Dict) -> str:
    processor = TableProcessor()
    return processor.get_table_summary(processed_tables)